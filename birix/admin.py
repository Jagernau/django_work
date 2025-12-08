from django.contrib import admin
from django.contrib import messages
from birix.sendmail import sendmailclient, sendmailmanager
from birix.models import *
from django.http import HttpResponse, HttpResponseRedirect, JsonResponse
from django.contrib.auth.mixins import LoginRequiredMixin
import openpyxl
from datetime import datetime, timedelta
from django.utils.html import format_html
import pytz
from django.urls import reverse, path
from django.utils.safestring import mark_safe
from birix.okdesk_funcs import create_okdesk_ticket 
from rangefilter.filters import DateRangeFilter
from django import forms
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.utils.translation import gettext_lazy as _

class ContragentsAdmin(LoginRequiredMixin, admin.ModelAdmin):

    actions = ['download_excel',]


    list_display = (
            "ca_name", 
            "ca_shortname",
            "ca_inn",
            "ca_kpp",
            "ca_field_of_activity",
            )
    list_filter = (
            "ca_type",
            "registration_date",
            "ca_field_of_activity",
            "key_manager",
            )
    search_fields = (
            "ca_name",
            "ca_inn",
            "ca_kpp",
            "ca_field_of_activity",
            )

    fieldsets = (
            (None, {
                'fields': (
                    'ca_name',
                    'ca_shortname',
                    'ca_inn',
                    'ca_kpp',
                    'ca_field_of_activity',
                )
            }),
        
        )
    list_per_page = 20

    def download_excel(self, request, queryset):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Contragents Data"

            # Write headers
            header_row = ["ca_name", "ca_shortname", "ca_inn", "ca_kpp", "ca_field_of_activity"]
            for col_num, header in enumerate(header_row, 1):
                worksheet.cell(row=1, column=col_num).value = header

            # Write data rows
            row_num = 2
            for contragent in queryset:
                worksheet.cell(row=row_num, column=1).value = contragent.ca_name
                worksheet.cell(row=row_num, column=2).value = contragent.ca_shortname
                worksheet.cell(row=row_num, column=3).value = contragent.ca_inn
                worksheet.cell(row=row_num, column=4).value = contragent.ca_kpp
                worksheet.cell(row=row_num, column=5).value = contragent.ca_field_of_activity
                row_num += 1

            # Set content type and attachment filename
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=contragents.xlsx'

            # Write workbook to response
            workbook.save(response)
            return response



class LoginUsersAdmin(LoginRequiredMixin,admin.ModelAdmin):

    actions = ['download_excel', 'send_access_mail_manager', 'send_access_mail_client']

    list_display = (
            "login",
            "password",
            "date_create",
            "system",
            "contragent",
            "comment_field",
            "account_status",
            "get_by_manager",
            "get_it_manager",
            )

    list_filter = (
            ('date_create', DateRangeFilter),
            "system",
            "date_create",
            "contragent__service_manager",
            "contragent__key_manager",
            "account_status",
            )
    search_fields = (
            "client_name",
            "login",
            "comment_field",
            "contragent__ca_name",
    )

    fieldsets = (
            (None, {
                'fields': (
                    'login',
                    'email',
                    'password',
                    'date_create',
                    'system',
                    'contragent',
                    'comment_field',
                    'account_status',
                )
            }),
    )

    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'login',
                    'email',
                    'password',
                    'date_create',
                    'system',
                    'contragent',
                    'comment_field',
                    'account_status',
                )
            })
    )
    autocomplete_fields = ('contragent',)
    list_per_page = 20

# Отправка сообщений с данными для входа по чекбоксам
    def send_access_mail_manager(self, request, queryset):
        for obj in queryset:
            try:
                manager_name = obj.contragent.key_manager
                contragent_name = obj.contragent.ca_name
                manager_email = CaContacts.objects.filter(ca_contact_surname = str(manager_name).split(' ')[0]).first().ca_contact_email
                system_url = obj.system.mon_url
            except Exception as e:
                messages.error(request, f'Ошибка, у контрагента не указан менеджер.{e}.')
            else:
                try:
                #Отправка менеджеру
                    sendmailmanager(manager_email, obj.login, obj.password, contragent_name, system_url, request.user.last_name)
                    obj.account_status = 2
                    obj.save()
                #Отправка создателю
                    sendmailmanager(request.user.email, obj.login, obj.password, contragent_name, system_url, request.user.last_name)
                #Отправка начальству
                    if request.user.username != 'alexandr_master':
                        sendmailmanager('it5@suntel-nn.ru', obj.login, obj.password, contragent_name, system_url, request.user.last_name)
                    messages.success(request, f'Письмо успешно отправлено менеджеру {manager_name} для {contragent_name}.')
                except Exception as e:
                    messages.error(request, f'Ошибка при отправке письма: {e}.')
        return None

# Отправка сообщений с данными для входа по чекбоксам
    def send_access_mail_client(self, request, queryset):
        for obj in queryset:
            try:
                contragent_name = obj.contragent.ca_name
                system_url = obj.system.mon_url
                mon_system_name = obj.system.mon_sys_name
            except Exception as e:
                messages.error(request, f'Ошибка, неправильно заполнена форма клиента (в 1с).{e}.')
            else:
                try:
                #Отправка клиенту
                    sendmailclient(obj.email, obj.login, obj.password, mon_system_name, system_url)
                #Отправка начальству
                    sendmailclient('it5@suntel-nn.ru', obj.login, obj.password, contragent_name, system_url)
                    obj.account_status = 2
                    obj.save()
                    messages.success(request, f'Письмо успешно отправлено для {contragent_name}.')
                except Exception as e:
                    messages.error(request, f'Ошибка при отправке письма: {e}.')
        return None
    

    def download_excel(self, request, queryset):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "LoginUsers Data"

            # Write headers
            header_row = ["login", "email", "password", "date_create", "system", "contragent", "comment_field", "account_status"]
            for col_num, header in enumerate(header_row, 1):
                worksheet.cell(row=1, column=col_num).value = header

            # Write data rows
            row_num = 2
            for loginuser in queryset:
                worksheet.cell(row=row_num, column=1).value = str(loginuser.login)
                worksheet.cell(row=row_num, column=2).value = str(loginuser.email)
                worksheet.cell(row=row_num, column=3).value = str(loginuser.password)
                worksheet.cell(row=row_num, column=4).value = str(loginuser.date_create)
                worksheet.cell(row=row_num, column=5).value = str(loginuser.system)
                worksheet.cell(row=row_num, column=6).value = str(loginuser.contragent)
                worksheet.cell(row=row_num, column=7).value = str(loginuser.comment_field)
                worksheet.cell(row=row_num, column=8).value = str(loginuser.account_status)
                row_num += 1

            # Set content type and attachment filename
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=loginusers.xlsx'

            # Write workbook to response
            workbook.save(response)
            return response


    def get_by_manager(self, obj):
        if obj.contragent_id != None:
            by_manager = Contragents.objects.filter(ca_id=obj.contragent_id).first().key_manager
            return by_manager

    def get_it_manager(self, obj):
        if obj.contragent_id != None:
            by_manager = Contragents.objects.filter(ca_id=obj.contragent_id).first().service_manager
            return by_manager

    get_by_manager.short_description = 'Менеджер по продажам'
    get_it_manager.short_description = 'ИТ специалист'

class CaObjectsAdmin(LoginRequiredMixin,admin.ModelAdmin):

    actions = ['download_excel',]
    readonly_fields = ('sys_mon', 'object_name', 'object_status', 'owner_contragent', 'owner_user', 'contragent', 'imei')


    list_display = (
            "sys_mon",
            "object_name",
            "object_status",
            "owner_contragent",
            "owner_user",
            "contragent",
            "imei",
            "get_device",
            "get_sim",
            "sys_mon_object_id",
            "upload_button",
            "view_file_button",
            )

    list_filter = (
            "object_status",
            "sys_mon",
#            "contragent",
#            "contragent",
            )
    search_fields = (
            "object_name",
            "contragent__ca_name",
#            "object_status__status_id",
            "owner_user",
            "owner_contragent",
            "imei",
    )

    fieldsets = (
            (None, {
                'fields': (
                    'sys_mon',
                    'object_name',
                    'object_status',
                    'owner_contragent',
                    'owner_user',
                    'contragent',
                    'imei',

                )
            }),
    )
    list_per_page = 20

    def get_device(self, obj):
        if Devices.objects.filter(device_imei=obj.imei).first():
            if obj.imei == None:
                return "Терминал не найден"
            if obj.imei == Devices.objects.filter(device_imei=obj.imei).first().device_imei:
                return [
                        Devices.objects.filter(device_imei=obj.imei).first().device_serial,
                        Devices.objects.filter(device_imei=obj.imei).first().devices_brand,
                        ]

    def get_sim(self, obj):
        if SimCards.objects.filter(terminal_imei=obj.imei).first():
            if obj.imei == None:
                return "Сим не найден"
                
            if obj.imei == SimCards.objects.filter(terminal_imei=obj.imei).first().terminal_imei:
                return [SimCards.objects.filter(terminal_imei=obj.imei).first().sim_iccid,
                        SimCards.objects.filter(terminal_imei=obj.imei).first().sim_tel_number,
                        SimCards.objects.filter(
                            terminal_imei=obj.imei
                            ).first().sim_cell_operator.name
                        ]

    get_device.short_description = 'Терминал'
    get_sim.short_description = 'Симкарта'


    def upload_button(self, obj):
        return mark_safe(f'<a class="button" href="{reverse("upload_file", args=[obj.id])}">Загрузить</a>')
    upload_button.short_description = 'Загрузить файл'


    def view_file_button(self, obj):
        """Кнопка для просмотра файла в Яндекс.Диске"""
        # Формируем базовое имя файла (без даты)
        base_name = f"{obj.object_name}".replace('/', '!')
        # Формируем URL для проверки файлов
        url = reverse('check_yandex_files', args=[obj.id])
        return mark_safe(
            f'<a class="button" href="{url}" style="background-color:#6c5ce7;color:white;">'
            f'🔍 Просмотреть'
            f'</a>'
        )
    view_file_button.short_description = 'Файлы на диске'

    def download_excel(self, request, queryset):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "CaObjects Data"

            # Write headers
            header_row = ["sys_mon", "object_name", "object_status", "owner_contragent", "owner_user", "contragent", "imei"]
            for col_num, header in enumerate(header_row, 1):
                worksheet.cell(row=1, column=col_num).value = header

            # Write data rows
            row_num = 2
            for caobject in queryset:
                worksheet.cell(row=row_num, column=1).value = str(caobject.sys_mon)
                worksheet.cell(row=row_num, column=2).value = str(caobject.object_name)
                worksheet.cell(row=row_num, column=3).value = str(caobject.object_status)
                worksheet.cell(row=row_num, column=4).value = str(caobject.owner_contragent)
                worksheet.cell(row=row_num, column=5).value = str(caobject.owner_user)
                worksheet.cell(row=row_num, column=6).value = str(caobject.contragent)
                worksheet.cell(row=row_num, column=7).value = str(caobject.imei)
                row_num += 1

            # Set content type and attachment filename
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=caobjects.xlsx'

            # Write workbook to response
            workbook.save(response)
            return response


class GlobalLogAdmin(LoginRequiredMixin,admin.ModelAdmin):
    list_display = (
            "section_type",
            "get_top_info",
            "get_obj_client",
            "field",
            "get_status_old",
            "get_status_new",
            "change_time",
            "sys_id",
            "action",
            )

    list_filter = (
            ('change_time', DateRangeFilter),
            "section_type",
            "sys_id",
            "field",
            "change_time",
            "action",
            )
    search_fields = (
            "section_type",
            "edit_id",
            "field",
            "change_time",
            "sys_id",
            "action",
            "new_value",
            "old_value",
            )
    fieldsets = (
            (None, {
                'fields': (
                    'section_type',
                    'edit_id',
                    'field',
                    'old_value',
                    'new_value',
                    'change_time',
                    'sys_id',
                    'action',
                )
            }),
    )
    list_per_page = 20
    date_hierarchy = 'change_time'


    def get_obj_client(self, obj):
        if obj.section_type == 'object':
            if CaObjects.objects.filter(id=obj.edit_id).first():
                return CaObjects.objects.filter(id=obj.edit_id).first().owner_contragent

        if obj.section_type == '1С_client':
            if Contragents.objects.filter(ca_id=obj.edit_id).first():
                return Contragents.objects.filter(ca_id=obj.edit_id).first().ca_name

    def get_status_old(self, obj):
        if obj.section_type == 'object' and obj.field == 'object_status_id':
            if obj.old_value == "0":
                return 'Не было'
            if obj.old_value == "1":
                return 'Новый не на абонентке'
            if obj.old_value == "2":
                return 'Тестоввый не на абонентке'
            if obj.old_value == "3":
                return 'На абонентке'
            if obj.old_value == "4":
                return 'Ждёт перевода'
            if obj.old_value == "5":
                return 'Приостановлен'
            if obj.old_value == "6":
                return 'Переведённый в другую систему'
            if obj.old_value == "7":
                return 'Деактивирован'
            else:
                return obj.old_value

        if obj.section_type == 'sim_card' and obj.field == 'status':
            if obj.old_value == "0":
                return 'Списана'
            if obj.old_value == "1":
                return 'Активна'
            if obj.old_value == "2":
                return 'Приостановлена'
            if obj.old_value == "3":
                return 'Первичная блокировка'
            if obj.old_value == "4":
                return 'Статус не известен'
            if obj.old_value == "5":
                return 'Сезонная блокировка'
            else:
                return obj.old_value
        else:
            return obj.old_value

    def get_status_new(self, obj):
        if obj.section_type == 'object' and obj.field == 'object_status_id':
            if obj.new_value == "0":
                return 'Не было'
            if obj.new_value == "1":
                return 'Новый не на абонентке'
            if obj.new_value == "2":
                return 'Тестоввый не на абонентке'
            if obj.new_value == "3":
                return 'На абонентке'
            if obj.new_value == "4":
                return 'Ждёт перевода'
            if obj.new_value == "5":
                return 'Приостановлен'
            if obj.new_value == "6":
                return 'Переведённый в другую систему'
            if obj.new_value == "7":
                return 'Деактивирован'
            else:
                return obj.new_value

        if obj.section_type == 'sim_card' and obj.field == 'status':
            if obj.new_value == "0":
                return 'Списана'
            if obj.new_value == "1":
                return 'Активна'
            if obj.new_value == "2":
                return 'Приостановлена'
            if obj.new_value == "3":
                return 'Первичная блокировка'
            if obj.new_value == "4":
                return 'Статус не известен'
            if obj.new_value == "5":
                return 'Сезонная блокировка'
            else:
                return obj.new_value
        else:
            return obj.new_value



    def get_top_info(self, obj):
        info_id = obj.edit_id
        section = obj.section_type

        if section == "sim_card":
            sim = SimCards.objects.filter(sim_id=info_id).first()
            if sim:
                return sim.sim_iccid 

        if section == "object":
            obj = CaObjects.objects.filter(id=info_id).first()
            if obj:
                return obj.object_name



    get_obj_client.short_description = "Контрагент"
    get_status_old.short_description = "Старое значение"
    get_status_new.short_description = "Новое значение"
    get_top_info.short_description = "Детализация"




class SimCardsAdmin(LoginRequiredMixin,admin.ModelAdmin):

    actions = ['download_excel',]

    list_display = (
            "sim_iccid",
            "sim_tel_number",
            "sim_cell_operator",
            "sim_date",
            "contragent",
            'itprogrammer',
            'status',
            'block_start',
            "get_end_date",
            "sim_owner",
            "terminal_imei",
            'get_device',
            )

    list_filter = (
            ('sim_date', DateRangeFilter),
            ('block_start', DateRangeFilter),
            "sim_cell_operator",
            "sim_owner",
            "sim_date",
            'itprogrammer',
            'status',
            "block_start",
            )
    search_fields = (
            "sim_iccid",
            "sim_tel_number",
            "client_name",
            "sim_cell_operator__name",
            "sim_owner",
            "sim_date",
            "contragent__ca_name",
            "terminal_imei",
            "block_start",
            )
    fieldsets = (
            (None, {
                'fields': (
                    'sim_iccid',
                    'sim_tel_number',
                    'client_name',
                    'sim_cell_operator',
                    'sim_owner',
                    'sim_date',
                    'contragent',
                    "terminal_imei",
                    'itprogrammer',
                    'status',
                )
            }),
    )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'sim_iccid',
                    'sim_tel_number',
#                    'client_name',
                    'sim_cell_operator',
                    'sim_owner',
                    'sim_date',
                    'contragent',
                    "terminal_imei",
                    'itprogrammer',
                    'status',

                )
            })
    )
    autocomplete_fields = (
        'contragent',
    )
    list_per_page = 20
    date_hierarchy = 'sim_date'

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.order_by('-sim_date')

    def get_device(self, obj):
        if Devices.objects.filter(device_imei=obj.terminal_imei).first():
            if obj.terminal_imei == Devices.objects.filter(device_imei=obj.terminal_imei).first().device_imei:
                return Devices.objects.filter(device_imei=obj.terminal_imei).first().device_serial

    def get_end_date(self, obj):
        if obj.block_start:
            # Получаем текущую дату с учетом временной зоны
            current_date = datetime.now(pytz.utc)  # Используем UTC или вашу локальную временную зону
            # Приводим block_start к UTC, если он offset-aware
            if obj.block_start.tzinfo is None:
                block_start = obj.block_start.replace(tzinfo=pytz.utc)  # Присваиваем временную зону
            else:
                block_start = obj.block_start

            # Вычисляем дату окончания блокировки
            end_date = block_start + timedelta(days=180)  # 180 дней = 6 месяцев
            
            # Проверяем, превышает ли дата окончания текущую дату
            if end_date < current_date:
                return format_html('<span style="color: red;">{}</span>', end_date.strftime('%Y-%m-%d'))
            
            return end_date.strftime('%Y-%m-%d')
        
        return None

    get_device.short_description = 'Сер. терм'
    get_end_date.short_description = 'Окончание блокировки'

    def download_excel(self, request, queryset):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "SimCards Data"

            # Write headers
            header_row = ["sim_iccid", "sim_tel_number", "client_name", "sim_cell_operator", "sim_owner", "sim_date", "contragent", "terminal_imei", "itprogrammer", "status", "block_start"]
            for col_num, header in enumerate(header_row, 1):
                worksheet.cell(row=1, column=col_num).value = header

            # Write data rows
            row_num = 2
            for sim in queryset:
                worksheet.cell(row=row_num, column=1).value = str(sim.sim_iccid)
                worksheet.cell(row=row_num, column=2).value = str(sim.sim_tel_number)
                worksheet.cell(row=row_num, column=3).value = str(sim.client_name)
                worksheet.cell(row=row_num, column=4).value = str(sim.sim_cell_operator)
                worksheet.cell(row=row_num, column=5).value = str(sim.sim_owner)
                worksheet.cell(row=row_num, column=6).value = str(sim.sim_date)
                worksheet.cell(row=row_num, column=7).value = str(sim.contragent)
                worksheet.cell(row=row_num, column=8).value = str(sim.terminal_imei)
                worksheet.cell(row=row_num, column=9).value = str(sim.itprogrammer)
                worksheet.cell(row=row_num, column=10).value = str(sim.status)
                worksheet.cell(row=row_num, column=10).value = str(sim.block_start)
                row_num += 1

            # Set content type and attachment filename
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=caobjects.xlsx'

            # Write workbook to response
            workbook.save(response)
            return response

class DevicesAdmin(LoginRequiredMixin,admin.ModelAdmin):

    actions = ['download_excel',]

    list_display = (
            "device_serial",
            "device_imei",
            "device_owner",
#            "client_name",
            "terminal_date",
            "print_button",
            "devices_brand",
            "sys_mon",
            "contragent",
            'itprogrammer',
            'get_sim',
            'coment',
            )

    list_filter = (
            ('terminal_date', DateRangeFilter),
            "devices_brand",
            "terminal_date",
            'itprogrammer',
            "devices_brand__devices_vendor",
            "sys_mon",
            "device_owner",
            'coment',
            )
    search_fields = (
            "device_serial",
            "device_imei",
            "client_name",
            "name_it",
            "contragent__ca_name",
            'coment',

    )
    fieldsets = (
            (None, {
                'fields': (
                    'device_serial',
                    'device_imei',
                    "device_owner",
                    'terminal_date',
                    'devices_brand',
                    'sys_mon',
                    'contragent',
                    'itprogrammer',
                    'coment',
                )
            }),
    )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'device_serial',
                    'device_imei',
                    "device_owner",
                    'terminal_date',
                    'devices_brand',
                    'sys_mon',
                    'contragent__ca_id',
                    'itprogrammer',
                    'coment',
                )
            })

    )
    #raw_id_fields = ['contragent']
    autocomplete_fields = (
        'contragent',
    )
    list_per_page = 20
    date_hierarchy = 'terminal_date'

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "devices_brand":
            kwargs["queryset"] = DevicesBrands.objects.all().order_by('name')
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('save_form/', self.admin_site.admin_view(self.save_form), name='save_form'),
        ]
        return custom_urls + urls

    def handle_save_form(self, request):
        if request.method == "POST":
            data = request.POST.dict()
            data.pop('device_serial', None)
            data.pop('device_imei', None)
            request.session['form_data'] = data
            return JsonResponse({'success': True})
        return JsonResponse({'success': False})

    def add_view(self, request, form_url='', extra_context=None):
        if request.method == "GET" and 'form_data' in request.session:
            form_data = request.session.pop('form_data')
            form = self.get_form(request)(initial=form_data)
        else:
            form = self.get_form(request)()
        
        return super().add_view(request, form_url, extra_context={'form': form})

    def print_button(self, obj):
        return mark_safe(f'<a class="button" href="{reverse("print", args=[obj.device_id])}">🖨</a>')
    print_button.short_description = ''


    def get_sim(self, obj):
        if SimCards.objects.filter(terminal_imei=obj.device_imei).first():
            if obj.device_imei == SimCards.objects.filter(terminal_imei=obj.device_imei).first().terminal_imei:
                return SimCards.objects.filter(terminal_imei=obj.device_imei).first().sim_iccid

    get_sim.short_description = 'Симкарта на устройстве'
#    list_display_links = ('get_sim',)

    def download_excel(self, request, queryset):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Terminal Data"

            # Write headers
            header_row = [
                    "device_serial",
                    "device_imei",
                    "client_name",
                    "terminal_date",
                    "devices_brand",
                    "sys_mon",
                    "contragent",
                    'itprogrammer',
                    ]
            for col_num, header in enumerate(header_row, 1):
                worksheet.cell(row=1, column=col_num).value = header

            # Write data rows
            row_num = 2
            for sim in queryset:
                worksheet.cell(row=row_num, column=1).value = str(sim.device_serial)
                worksheet.cell(row=row_num, column=2).value = str(sim.device_imei)
                worksheet.cell(row=row_num, column=3).value = str(sim.client_name)
                worksheet.cell(row=row_num, column=4).value = str(sim.terminal_date)
                worksheet.cell(row=row_num, column=5).value = str(sim.devices_brand)
                worksheet.cell(row=row_num, column=6).value = str(sim.sys_mon)
                worksheet.cell(row=row_num, column=7).value = str(sim.contragent)
                worksheet.cell(row=row_num, column=8).value = str(sim.itprogrammer)
                row_num += 1

            # Set content type and attachment filename
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=terminal.xlsx'

            # Write workbook to response
            workbook.save(response)
            return response

class DeviceBrandsAdmin(LoginRequiredMixin,admin.ModelAdmin):
    list_display = (
            "name",
            "devices_vendor",
            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'name',
                    'devices_vendor',

                )
            })
    )

    list_filter = (
            "devices_vendor",
    )
    search_fields = (
            "name",
            "devices_vendor__vendor_name",
    )

class ContactsAdmin(admin.ModelAdmin):

    actions = ['copy_record']

    def copy_record(self, request, queryset):
        for obj in queryset:
            obj.id = None
            obj.save()

    copy_record.short_description = "Копировать запись"

    list_display = (
            "ca_contact_cell_num",
            "ca_contact_email",
            "ca_contact_name",
            "ca_contact_surname",
            "ca",
            "ca_contact_position",
            )

    fieldsets = (
            (None, {
                'fields': (
                    'ca_contact_cell_num',
                    'ca_contact_email',
                    'ca_contact_name',
                    'ca_contact_surname',
                    'ca',
                    'ca_contact_position',

                )
            }),
    )

    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'ca_contact_cell_num',
                    'ca_contact_email',
                    'ca_contact_name',
                    'ca_contact_surname',
                    'ca',
                    'ca_contact_position',
                )
            })
    )

    search_fields = (
            "ca_contact_cell_num",
            "ca_contact_email"
            "ca_contact_name",
            "ca_contact_surname",
            "ca",
    )
    list_filter = (
            "ca_contact_position",
    )
    autocomplete_fields = (
        'ca',
    )
    
class DevicesCommandAdmin(admin.ModelAdmin):

    actions = ['copy_record']

    def copy_record(self, request, queryset):
        for obj in queryset:
            obj.id = None
            obj.save()

    copy_record.short_description = "Копировать запись"

    list_display = (
            "command",
            "device_brand",
            "method",
            "description",
            )

    list_filter = (
            "device_brand",
            "method",
    )

    search_fields = (
            "command",
            "description",
    )

    fieldsets = (
            (None, {
                'fields': (
                    'command',
                    'device_brand',
                    'method',
                    'description',
                )
            }),
    )

    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'command',
                    'device_brand',
                    'method',
                    'description',
                )
            })
    )


class LogAdmin(admin.ModelAdmin):
    list_display = (
            "action_time",
            "object_id",
            "object_repr",
            "get_change_message",
            "get_change_info",
            "content_type",
            "user",
            )
    list_filter = (
            ('action_time', DateRangeFilter),
            )

    def get_change_message(self, obj):
        if obj.action_flag == 1:
            return 'Добавлен новый объект'
        if obj.action_flag == 2:
            return 'Объект изменен'
        if obj.action_flag == 3:
            return 'Объект удален'

    def get_change_info(self, obj):
        message = str(obj.change_message).replace('[', '').replace(']', '').replace('{"changed": {"', "").replace('{"added": {}}', "").replace('"}}', "").replace('fields": "', "")
        clear_message = message.encode('utf-8').decode('unicode_escape')
        return clear_message

    get_change_info.short_description = 'Изменения'

    get_change_message.short_description = 'Действие'


class DeviceVendorAdmin(admin.ModelAdmin):
    list_display = (
            "vendor_name",
            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'vendor_name',

                )
            })
    )
class MonitoringSystemAdmin(admin.ModelAdmin):
    list_display = (
            "mon_sys_name",
            "mon_sys_ca_obj_price_default",
            "mon_sys_obj_price_suntel",
            "mon_url",

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'vendor_name',
                    "mon_sys_ca_obj_price_default",
                    "mon_sys_obj_price_suntel",
                    'mon_url',

                )
            })
    )

class ObjectRetranslatorsAdmin(admin.ModelAdmin):
    list_display = (
            "retranslator_name",
            "retranslator_suntel_price",
            "retranslator_ca_price",
            "retrans_adres",
            "retrans_protocol",

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    'retranslator_name',
                    "retranslator_suntel_price",
                    "retranslator_ca_price",
                    "retrans_adres",
                    "retrans_protocol",

                )
            })
    )
class GroupObjectRetransAdmin(admin.ModelAdmin):
    list_display = (
            "obj",
            "retr",
            "client_name"

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "obj__id",
                    "retr",

                )
            })
    )
    list_filter = (
            "retr",
            )


    autocomplete_fields = (
        'obj',
    )

    search_fields = (
            "obj__object_name",
            "obj__contragent_id__ca_name"
            )


class ObjectSensorsAdmin(admin.ModelAdmin):
    list_display = (
            "sensor_type",
            "sensor_model",
            "sensor_technology",
            "sensor_connect_type",
            "client",
            "sensor_serial",
            "name_installer",
            "installer_id",
            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "sensor_type",
                    "sensor_model",
                    "sensor_technology",
                    "sensor_connect_type",
                    "client__ca_id",
                    "sensor_serial",
                    "name_installer",
                    "installer_id",

                )
            })
    )
    list_filter = (
            "sensor_type",
            "sensor_model",
            "sensor_technology",
            "sensor_connect_type",
    )
    search_fields = (
            "sensor_serial",
            "sensor_type",
            "sensor_technology",
    )
    autocomplete_fields = (
        'client',
    )


class WarehouseAdmin(admin.ModelAdmin):
    list_display = (
            "add_date",
            "serial_number",
            "availability",
            "terminal_model",
            "sensor",
            "delivery_date",
            "client",
            "comment",
            "whom_issued",
            "affiliation",
            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "serial_number",
                    "availability",
                    "terminal_model__id",
                    "sensor__sensor_id",
                    "delivery_date",
                    "client__contragent_id",
                    "comment",
                    "whom_issued",
                    "affiliation",
                )
            })
    )
    list_filter = (
            "affiliation",
            "delivery_date",
            "availability",
            "terminal_model",
            "sensor",
    )
    autocomplete_fields = (
        'terminal_model',
        'sensor',
        'client',
    )

class SensorBrandsAdmin(admin.ModelAdmin):
    list_display = (
            "name",
            "sensor_vendor",

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "sensor_vendor__id",
                    "name",

                )
            })
    )
    list_filter = (
            'sensor_vendor__id',
            )

    search_fields = (
            "name",
            "sensor_vendor__vendor_name",
    )


    autocomplete_fields = (
        'sensor_vendor',
    )



class SensorVendorAdmin(admin.ModelAdmin):
    list_display = (
            "name",

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "name",

                )
            })
    )
    search_fields = (
            "name",
    )

"""class DeviceDiagnosicAdmin(admin.ModelAdmin):
    list_display = (
            "device",
            "get_imei",
            "get_klient",
            "programmer",
            "brought",
            "comment",
            "accept_date",
            "transfer_date",
            "whom_tranfer",

            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "device",
                    "programmer",
                    "brought",
                    "comment",
                    "transfer_date",
                    "whom_tranfer",

                )
            }),)
    
    list_filter = (
            ('accept_date', DateRangeFilter),
            ('transfer_date', DateRangeFilter),
            'programmer',
            'brought',
            'whom_tranfer',
            )
    search_fields = (
            "comment",
            "device__device_serial",
            "device__contragent__ca_name",
            "device__device_imei",
    )
    autocomplete_fields = (
        'device',
    )
    date_hierarchy = 'transfer_date'"""

class DeviceDiagnosticForm(forms.ModelForm):
    class Meta:
        model = DevicesDiagnostics
        fields = '__all__'
    
    def init(self, *args, **kwargs):
        super().init(*args, **kwargs)
        # Группировка полей для лучшего отображения
        for field_name in self.fields:
            if field_name.endswith('_comment'):
                self.fields[field_name].widget.attrs.update({
                    'style': 'width: 250px; margin-left: 10px; vertical-align: middle;'
                })
            elif field_name.endswith('_check'):
                self.fields[field_name].widget.attrs.update({
                    'style': 'vertical-align: middle; margin-right: 5px;'
                })


class DeviceDiagnosicAdmin(admin.ModelAdmin):
    form = DeviceDiagnosticForm
    
    list_display = (
        "device",
        "get_imei",
        "get_klient",
        "programmer",
        "print_button",
        "brought",
        "comment",
        "accept_date",
        "transfer_date",
        "whom_tranfer",
        
    )

    def print_button(self, obj):
        if obj.id:
            # Используем имя URL из urls.py
            url = reverse('print_diagnostic', args=[obj.id])
            return format_html(
                '<a href="{}" target="_blank" class="button" style="background-color: #4CAF50; color: white; padding: 5px 10px; text-decoration: none; border-radius: 3px;">🖨</a>',
                url
            )
        return "-"
    print_button.short_description = "Печать"
    print_button.allow_tags = True

    # Определяем поля для разных разделов формы
    fieldsets = (
        ('Основная информация', {
            'fields': (
                'device',
                'programmer',
                'brought',
                'whom_tranfer',
            )
        }),
        ('Даты', {
            'fields': (
                ('accept_date', 'transfer_date'),
            )
        }),
        ('Диагностика', {
            'fields': (
                ('USB_comment', 'USB_check'),
                ('PWR_comment', 'PWR_check'),
                ('PWR_AKB_comment', 'PWR_AKB_check'),
                ('FIRMWARE_comment', 'FIRMWARE_check'),
                ('SATS_comment', 'SATS_check'),
                ('GSM_comment', 'GSM_check'),
                ('ONLINE_comment', 'ONLINE_check'),
                ('P485_comment', 'P485_check'),
                ('DIGIT_PORT_comment', 'DIGIT_PORT_check'),
                ('ANALOG_PORT_comment', 'ANALOG_PORT_check'),
            )
        }),
        ('Комплектация', {
            'fields': (
                'GSM_antenna_check',
                'GPS_antenna_check',
                'CABEL_check',
                ('SIM_check', 'SIM_comment'),
            )
        }),
        ('', {
            'fields': (
                'comment',
            )
        }),
    )
    
    # Для формы создания используем ту же структуру
    add_fieldsets = (
        ('Основная информация', {
            'fields': (
                'device',
                'programmer',
                'brought',
                'whom_tranfer',
            )
        }),
        ('Даты', {
            'fields': (
                ('accept_date', 'transfer_date'),
            )
        }),
        ('Диагностика', {
            'fields': (
                ('USB_comment', 'USB_check'),
                ('PWR_comment', 'PWR_check'),
                ('PWR_AKB_comment', 'PWR_AKB_check'),
                ('FIRMWARE_comment', 'FIRMWARE_check'),
                ('SATS_comment', 'SATS_check'),
                ('GSM_comment', 'GSM_check'),
                ('ONLINE_comment', 'ONLINE_check'),
                ('P485_comment', 'P485_check'),
                ('DIGIT_PORT_comment', 'DIGIT_PORT_check'),
                ('ANALOG_PORT_comment', 'ANALOG_PORT_check'),
            )
        }),
        ('Комплектация', {
            'fields': (
                'GSM_antenna_check',
                'GPS_antenna_check',
                'CABEL_check',
                ('SIM_check', 'SIM_comment'),
            )
        }),
        ('',{
            'fields': (
                'comment',
            )
        }),
        )

    def change_view(self, request, object_id, form_url='', extra_context=None):
        """Добавляем кнопку печати в контекст"""
        extra_context = extra_context or {}
        obj = self.get_object(request, object_id)
        if obj:
            extra_context['print_url'] = reverse('print_diagnostic', args=[object_id])
        return super().change_view(request, object_id, form_url, extra_context)

    def get_fieldsets(self, request, obj=None):
        if not obj:
            return self.add_fieldsets
        return super().get_fieldsets(request, obj)
    
    list_filter = (
        ('accept_date', DateRangeFilter),
        ('transfer_date', DateRangeFilter),
        'programmer',
        'brought',
        'whom_tranfer',
    )
    
    search_fields = (
        "comment",
        "device__device_serial",
        "device__contragent__ca_name",
        "device__device_imei",
    )
    
    autocomplete_fields = ('device',)
    date_hierarchy = 'transfer_date'




    def get_klient(self, obj):
        client = obj.device.contragent
        if client:
            return client.ca_name
        else:
            return "NONE"

    get_klient.short_description = 'Клиент'

    def get_imei(self, obj):
        return obj.device.device_imei

    get_imei.short_description = 'IMEI'

    def save_model(self, request, obj, form, change):
        """Отправляем данные в Okdesk, если whom_tranfer стало 1 при создании или изменении."""

        
        device = Devices.objects.filter(device_id=obj.device_id).first()
        if device.contragent == None:
            messages.error(request, "Не возможно создать диагностику, тк у терминала не прописан КЛИЕНТ")
            return

        previous_obj = None
        is_new = not obj.pk  # Проверяем, новый объект или нет

        if not is_new:
            previous_obj = DevicesDiagnostics.objects.get(pk=obj.pk)

        super().save_model(request, obj, form, change)

        # Отправляем данные в Okdesk:
        # - если объект новый и whom_tranfer = 1
        # - если объект существовал и whom_tranfer изменился с другого значения на 1
        if (is_new and obj.whom_tranfer == 1) or (previous_obj and previous_obj.whom_tranfer != 1 and obj.whom_tranfer == 1):
            device = Devices.objects.filter(device_id=obj.device_id).first()

            if device:
                imei = device.device_imei
                sm_object_abon = CaObjects.objects.filter(imei=imei).filter(object_status=3).first()

                if sm_object_abon:
                    obj_ok_id = sm_object_abon.ok_desk_id
                    if obj_ok_id:
                        ok_client_id = sm_object_abon.contragent.ok_desk_id
                        if ok_client_id:
                            pb_title = "Приостановить объект(АТ в ремонт)"
                            pb_desc = f"IMEI терминала: <b>{imei}</b>"
                            result_create = create_okdesk_ticket(
                                    object_name=sm_object_abon, 
                                    owner=ok_client_id, 
                                    object_id=obj_ok_id, 
                                    employ=obj.programmer.last_name, 
                                    problem_title=pb_title, 
                                    problem_desc=pb_desc,
                                    type_req="inner_proist_remont"
                                    )
                            if result_create[1] == False:
                                messages.error(request, result_create[0])
                            else:
                                messages.success(request, result_create[0])

                else:
                    messages.success(request, f"Объект в СМ не найден, или не на абонентке")


        if (is_new and obj.whom_tranfer == 0 and obj.brought == 1) or (previous_obj and previous_obj.whom_tranfer != 0 and obj.whom_tranfer == 0 and obj.brought == 1):
            device = Devices.objects.filter(device_id=obj.device_id).first()

            if device:
                imei = device.device_imei
                sm_object_abon = CaObjects.objects.filter(imei=imei).exclude(object_status=3).first()

                if sm_object_abon:
                    obj_ok_id = sm_object_abon.ok_desk_id
                    if obj_ok_id:
                        ok_client_id = sm_object_abon.contragent.ok_desk_id
                        if ok_client_id:
                            pb_title = "Активировать объект(АТ из ремонта)"
                            pb_desc = f"IMEI терминала: <b>{imei}</b>"
                            result_create = create_okdesk_ticket(
                                    object_name=sm_object_abon, 
                                    owner=ok_client_id, 
                                    object_id=obj_ok_id, 
                                    employ=obj.programmer.last_name, 
                                    problem_title=pb_title, 
                                    problem_desc=pb_desc,
                                    type_req="inner_vosst_obj_posle_remonta"
                                    )
                            if result_create[1] == False:
                                messages.error(request, result_create[0])
                            else:
                                messages.success(request, result_create[0])

                else:
                    messages.success(request, f"Объект в СМ не найден, или не снимался с абонентки")

class OnecContractsAdmin(admin.ModelAdmin):
    list_display = (
            "name_contract",
            "contract_number",
            "contract_date",
            "contract_status",
            "organization",
            "counterparty",
            "contract_purpose",
            "type_calculations",
            "category",
            "manager",
            "subdivision",
            "contact_person",
            "detailed_calculations",
            "ok_desk_id",
            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "name_contract",
                    "contract_number",
                    "contract_date",
                    "contract_status",
                    "organization",
                    "partner",
                    "counterparty",
                    "contract_commencement_date",
                    "contract_expiration_date",
                    "contract_purpose",
                    "type_calculations",
                    "category",
                    "manager",
                    "subdivision",
                    "contact_person",
                    "organization_bank_account",
                    "counterparty_bank_account",
                    "detailed_calculations",
                    "unique_partner_identifier",
                    "unique_counterparty_identifier",
                    "ok_desk_id",
                )
            })
    )
    list_filter = (
            ('contract_date', DateRangeFilter),

            "contract_date",
            "contract_status",
 
            "contract_purpose",
            "type_calculations",
            "category",
            "manager",
            )
    search_fields = (
            "name_contract",
            "contract_number",
            "contract_status",
            "organization",
            "counterparty",
            "manager",
            "subdivision",
            "contact_person",
            "organization_bank_account",
            "counterparty_bank_account",
            "detailed_calculations",
    )


class ClientFilter(admin.SimpleListFilter):
    title = 'Клиент'
    parameter_name = 'client'

    def lookups(self, request, model_admin):
        # Получаем текущий queryset с учетом всех фильтров
        queryset = model_admin.get_queryset(request)
        
        # Получаем уникальные логины из текущей выборки
        logins = queryset.values_list('sys_login', flat=True).distinct()
        
        # Получаем связанных контрагентов
        clients = Contragents.objects.filter(
            loginusers__login__in=logins
        ).distinct().values_list('ca_id', 'ca_name')
        
        return [(client[0], client[1]) for client in clients]

    def queryset(self, request, queryset):
        if self.value():
            # Фильтруем по выбранному клиенту
            return queryset.filter(
                sys_login__in=LoginUsers.objects.filter(
                    contragent__ca_id=self.value()
                ).values_list('login', flat=True)
            )
        return queryset

class SpecialistFilter(admin.SimpleListFilter):
    title = 'Специалист'
    parameter_name = 'specialist'

    def lookups(self, request, model_admin):
        # Получаем текущий queryset
        queryset = model_admin.get_queryset(request)
        
        # Получаем логины из текущей выборки
        logins = queryset.values_list('sys_login', flat=True).distinct()
        
        # Получаем уникальных специалистов
        specialists = Contragents.objects.filter(
            loginusers__login__in=logins
        ).exclude(service_manager__exact='').values_list(
            'service_manager', flat=True
        ).distinct().order_by('service_manager')
        
        return [(s, s) for s in specialists if s]

    def queryset(self, request, queryset):
        value = self.value()
        if value:
            return queryset.filter(
                sys_login__in=LoginUsers.objects.filter(
                    contragent__service_manager=value
                ).values_list('login', flat=True)
            )
        return queryset


class InfoServObjAdmin(admin.ModelAdmin):
    actions = ['download_excel']
    list_display = (
        "client_name_display",  # Добавляем имя клиента
        "serv_obj_sys_mon",
        "info_obj_serv",
        "subscription_start_date",
        "subscription_end_date",
        "service_counter",
        "stealth_type",
        "monitoring_sys",
        "sys_login",
        "sys_password",
        "send_meth",
        "specialist_display",
    )
    
    search_fields = (
        "sys_login",  # Поиск по имени клиента
        "serv_obj_sys_mon__object_name",
        "sys_id_obj",
        "sys_login",
        "sys_password",
    )

    def subscription_start_date(self, obj):
        return obj.subscription_start.strftime("%d.%m.%Y") if obj.subscription_start else ""
    subscription_start_date.short_description = "Начало подписки"
    subscription_start_date.admin_order_field = 'subscription_start'

    def subscription_end_date(self, obj):
        return obj.subscription_end.strftime("%d.%m.%Y") if obj.subscription_end else ""
    subscription_end_date.short_description = "Окончание подписки"
    subscription_end_date.admin_order_field = 'subscription_end'

    def specialist_display(self, obj):
        """Безопасное получение специалиста с обработкой дубликатов"""
        users = LoginUsers.objects.filter(login=obj.sys_login)
        if users.exists():
            # Берем первого пользователя (можно добавить сортировку при необходимости)
            return users.first().contragent.service_manager or "Не указан"
        return "Не указан"
    specialist_display.short_description = "Специалист"

    
    def client_name_display(self, obj):
        """Безопасное получение имени клиента с обработкой дубликатов"""
        users = LoginUsers.objects.filter(login=obj.sys_login)
        if users.exists():
            # Берем первого пользователя
            return users.first().contragent.ca_name
        return "Клиент не найден"
    client_name_display.short_description = "Клиент"
    
    def get_queryset(self, request):
        """Оптимизация запросов с prefetch_related"""
        return super().get_queryset(request).select_related(
            'serv_obj_sys_mon',
            'info_obj_serv',
            'monitoring_sys'
        )

    # Остальные части кода остаются без изменений
    add_fieldsets = (
        (None, {
            'classes': ('wide',),
            'fields': (
                "serv_obj_sys_mon",
                "info_obj_serv",
                "subscription_start",
                "subscription_end",
                "service_counter",
                "stealth_type",
                "monitoring_sys",
                "sys_login",
                "sys_password",
            )
        })
    )
    
    list_filter = (
        SpecialistFilter,
        ClientFilter,
        ('subscription_start', DateRangeFilter),
        "info_obj_serv",
        "subscription_start",
        "subscription_end",
        "service_counter",
        "stealth_type",
        "monitoring_sys",
        "sys_id_obj",
        "sys_login",
    )
    
    date_hierarchy = 'subscription_start'
    autocomplete_fields = ('serv_obj_sys_mon',)
    exclude = ['sys_id_obj', 'monitoring_sys', "tel_num_user"]



    def save_model(self, request, obj, form, change):
        # Получаем связанный объект и тариф
        ca_object = obj.serv_obj_sys_mon
        
        try:
            tarif = InfoServTarifs.objects.get(tarif_id=obj.stealth_type)
        except InfoServTarifs.DoesNotExist:
            messages.error(request,"Выбранный тарифный план не существует!")
            return

        # Проверка пересечения временных интервалов
        base_q = Q(serv_obj_sys_mon=ca_object) & ~Q(pk=obj.pk)
        
        # Для объектов с указанной датой окончания
        if obj.subscription_end:
            overlap_q = Q(
                Q(subscription_start__lt=obj.subscription_end,
                subscription_end__gt=obj.subscription_start) |
                Q(subscription_end__isnull=True,
                subscription_start__lt=obj.subscription_end)
            )
        # Для бессрочных подписок
        else:
            overlap_q = Q(
                Q(subscription_end__gt=obj.subscription_start) |
                Q(subscription_end__isnull=True)
            )

        # Получаем пересекающиеся сервисы
        active_services = InfoServObj.objects.filter(base_q & overlap_q)
        active_count = active_services.count()

        # Проверка лимита тарифа
        if active_count >= tarif.count:
            messages.error(request,f"Превышен лимит тарифа {tarif.name}! Максимум: {tarif.count} Текущее количество: {active_count}")
            return
        # Проверка дат
        if obj.subscription_end and obj.subscription_end <= obj.subscription_start:
            messages.error(request,"Дата окончания должна быть позже даты начала!")
            return

        # Сохраняем системные данные
        obj.monitoring_sys = ca_object.sys_mon
        obj.sys_id_obj = ca_object.sys_mon_object_id

        try:
            super().save_model(request, obj, form, change)
        except Exception as e:
            messages.error(request,f"Ошибка сохранения: {e}")

    def download_excel(self, request, queryset):
        import openpyxl
        from django.http import HttpResponse
        from django.utils import timezone
        from openpyxl.utils import get_column_letter

        # Оптимизация запросов
        queryset = queryset.select_related(
            'serv_obj_sys_mon',
            'info_obj_serv',
            'monitoring_sys'
        )

        # Собираем все необходимые логины
        logins = list(queryset.values_list('sys_login', flat=True).distinct())
        
        # Получаем данные клиентов и специалистов
        login_users = LoginUsers.objects.filter(
            login__in=logins
        ).select_related('contragent')
        
        login_info = {
            lu.login: {
                'client': lu.contragent.ca_name if lu.contragent else "Не определено",
                'specialist': lu.contragent.service_manager if lu.contragent else "Не указан"
            }
            for lu in login_users
        }

        # Создаем книгу Excel
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Информационные сервисы"

        # Заголовки
        headers = [
            "Клиент",
            "Объект мониторинга",
            "Сервис",
            "Начало подписки",
            "Конец подписки",
            "Периодичность",
            "Автоматизм",
            "Система",
            "Логин",
            "Пароль",
            "Способ отправки",
            "Специалист"
        ]

        # Записываем заголовки
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)

        # Заполняем данные
        for row_num, obj in enumerate(queryset, start=2):
            client_data = login_info.get(obj.sys_login, {
                'client': "Не определено",
                'specialist': "Не указан"
            })

            # Форматируем даты
            start_date = obj.subscription_start.strftime("%d.%m.%Y") if obj.subscription_start else ""
            end_date = obj.subscription_end.strftime("%d.%m.%Y") if obj.subscription_end else ""

            row = [
                client_data['client'],
                str(obj.serv_obj_sys_mon),
                str(obj.info_obj_serv),
                start_date,
                end_date,
                obj.get_service_counter_display(),
                obj.get_stealth_type_display(),
                str(obj.monitoring_sys),
                obj.sys_login,
                obj.sys_password,
                obj.get_send_meth_display(),
                client_data['specialist']
            ]

            for col_num, value in enumerate(row, 1):
                worksheet.cell(row=row_num, column=col_num, value=value)

        # Настраиваем ширину столбцов
        column_widths = [30, 25, 30, 15, 15, 25, 25, 25, 20, 20, 20, 25]
        for i, width in enumerate(column_widths, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width

        # Формируем ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"export_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        workbook.save(response)
        return response

    def get_queryset(self, request):
        """Оптимизация запросов"""
        return super().get_queryset(request).select_related(
            'serv_obj_sys_mon',
            'info_obj_serv',
            'monitoring_sys'
        ).prefetch_related(
            models.Prefetch(
                'serv_obj_sys_mon',
                queryset=CaObjects.objects.select_related('sys_mon')
            )
        )

    download_excel.short_description = "Экспорт в Excel"


class InfoServTarifsAdmin(admin.ModelAdmin):
    list_display = (
            "name",
            "price",
            "count",
            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "name",
                    "price",
                    "count",
                )
            })
    )


class InfoServTarifClientAdmin(admin.ModelAdmin):
    list_display = (
            "tarif",
            "client",
            "start_tarif",
            "end_tarif",
            )
    add_fieldsets = (
            (None, {
                'classes': ('wide',),
                'fields': (
                    "tarif",
                    "client",
                    "start_tarif",
                    "end_tarif",
                )
            })
    )

    list_filter = (
                    "tarif",
                    "start_tarif",
                    "end_tarif",
            )
    search_fields = (
                    "tarif__name",
                    "client__ca_name",
    )
    # date_hierarchy = 'subscription_start'
    autocomplete_fields = (
        'client',
    )



class OnecContactsAdmin(admin.ModelAdmin):
    list_display = (
            "surname",
            "name",
            "position",
            "phone",
            "mobiletelephone",
            "email",
            "get_client",
            )
    list_filter = (
            "surname",
            "name",
            "position",
            )
    search_fields = (
            "surname",
            "name",
            "position",
            "phone",
            "mobiletelephone",
            "email",
    )
    readonly_fields = [field.name for field in OnecContacts._meta.fields]
    def has_add_permission(self, request):
        return False  # Отключает возможность добавления новых записей

    def get_client(self, obj):
        client_uid = obj.unique_partner_identifier

        if client_uid:
            client = Contragents.objects.filter(unique_onec_id=client_uid).first()
            if client:
                return client.ca_name

    get_client.short_description = "Контрагент"

    def get_search_results(self, request, queryset, search_term):
        # Стандартный поиск по заданным полям
        queryset, use_distinct = super().get_search_results(request, queryset, search_term)
        
        # Дополнительный поиск по имени контрагента в Contragents
        if search_term:
            # Ищем контрагентов по ca_name
            contragents = Contragents.objects.filter(ca_name__icontains=search_term)
            # Получаем их unique_onec_id
            onec_ids = contragents.values_list('unique_onec_id', flat=True)
            # Добавляем записи OnecContacts, связанные с найденными контрагентами
            queryset |= self.model.objects.filter(unique_partner_identifier__in=onec_ids)
        
        return queryset, use_distinct


class CellOperatorAdmin(admin.ModelAdmin):
    list_display = (
            "name",
            "ca_price",
            "sun_price",
            )
    list_filter = (
            "name",
            "ca_price",
            "sun_price",
            )


class BillingAdmin(admin.ModelAdmin):
    actions = ['download_excel',]
    list_display = (
            "record_time",
            "obj_name",
            "obj_status",
            "obj_status_name",
            "obj_group_name",
            "sys_mon_name",
            "obj_imei",
            "client_name",
            "client_inn",
            "client_kpp",
            "client_login",
            "sim_operat_name",
            "retrans_name",
            "sys_mon_price",
            "sim_price",
            "retrans_price",
            "total_sum",
            )
    list_filter = (
            ('record_time', DateRangeFilter),
            "obj_status",
            "obj_status_name",
            "sys_mon_name",
            "sim_operat_name",
            "retrans_name",
            )
    search_fields = (
            "obj_name",
            "obj_status_name",
            "obj_group_name",
            "obj_imei",
            "client_name",
            "client_kpp",
            "client_login",
            "retrans_name",
    )
    readonly_fields = [
            "obj_group_name",
            "obj_status_name",
            "obj_imei",
            "client_name",
            "client_inn",
            "client_kpp",
            "client_login",
            "retrans_name",
            ]
    def has_add_permission(self, request):
        return False  # Отключает возможность добавления новых записей


    def download_excel(self, request, queryset):
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Billing_data"

            # Write headers
            header_row = [
                "record_time",
                "obj_name",
                "obj_status",
                "obj_status_name",
                "obj_group_name",
                "sys_mon_name",
                "obj_imei",
                "client_name",
                "client_inn",
                "client_kpp",
                "client_login",
                "sim_operat_name",
                "retrans_name",
                "sys_mon_price",
                "sim_price",
                "retrans_price",
                "total_sum",
                    ]
            for col_num, header in enumerate(header_row, 1):
                worksheet.cell(row=1, column=col_num).value = header

            # Write data rows
            row_num = 2
            for bill in queryset:
                worksheet.cell(row=row_num, column=1).value = str(bill.record_time)
                worksheet.cell(row=row_num, column=2).value = str(bill.obj_name)
                worksheet.cell(row=row_num, column=3).value = str(bill.obj_status)
                worksheet.cell(row=row_num, column=4).value = str(bill.obj_status_name)
                worksheet.cell(row=row_num, column=5).value = str(bill.obj_group_name)
                worksheet.cell(row=row_num, column=6).value = str(bill.sys_mon_name)
                worksheet.cell(row=row_num, column=7).value = str(bill.obj_imei)
                worksheet.cell(row=row_num, column=8).value = str(bill.client_name)
                worksheet.cell(row=row_num, column=9).value = str(bill.client_inn)
                worksheet.cell(row=row_num, column=10).value = str(bill.client_kpp)
                worksheet.cell(row=row_num, column=11).value = str(bill.client_login)
                worksheet.cell(row=row_num, column=12).value = str(bill.sim_operat_name)
                worksheet.cell(row=row_num, column=13).value = str(bill.retrans_name)
                worksheet.cell(row=row_num, column=14).value = str(bill.sys_mon_price)
                worksheet.cell(row=row_num, column=15).value = str(bill.sim_price)
                worksheet.cell(row=row_num, column=16).value = str(bill.retrans_price)
                worksheet.cell(row=row_num, column=17).value = str(bill.total_sum)
                row_num += 1

            # Set content type and attachment filename
            response = HttpResponse(content_type='application/vnd.ms-excel')
            response['Content-Disposition'] = 'attachment; filename=terminal.xlsx'

            # Write workbook to response
            workbook.save(response)
            return response



class TransferClientAdmin(admin.ModelAdmin):
    list_display = (
            "client",
            "tranfer_sys_mon",
            "current_sys_mon",
            "start_trans",
            "end_trans",
            "terminal_command",
            "command_type",
            "device_model",
            )
    # list_filter = (
    #         "surname",
    #         "name",
    #         "position",
    #         )
    # search_fields = (
    #         "surname",
    #         "name",
    #         "position",
    #         "phone",
    #         "mobiletelephone",
    #         "email",
    # )
    autocomplete_fields = (
        'client',
    )

admin.site.register(Contragents, ContragentsAdmin)
admin.site.register(LoginUsers, LoginUsersAdmin)
admin.site.register(GlobalLogging, GlobalLogAdmin)
admin.site.register(CaObjects, CaObjectsAdmin)
admin.site.register(SimCards, SimCardsAdmin)
admin.site.register(Devices, DevicesAdmin)
admin.site.register(DevicesBrands, DeviceBrandsAdmin)
admin.site.register(CaContacts, ContactsAdmin)
admin.site.register(DevicesCommands, DevicesCommandAdmin)
admin.site.register(DjangoAdminLog, LogAdmin)
admin.site.register(DevicesVendor, DeviceVendorAdmin)
admin.site.register(MonitoringSystem, MonitoringSystemAdmin)
admin.site.register(ObjectRetranslators, ObjectRetranslatorsAdmin)
admin.site.register(GroupObjectRetrans, GroupObjectRetransAdmin)
#admin.site.register(ObjectSensors, ObjectSensorsAdmin)
#admin.site.register(EquipmentWarehouse, WarehouseAdmin)
#admin.site.register(SensorBrands, SensorBrandsAdmin)
#admin.site.register(SensorVendor, SensorVendorAdmin)
admin.site.register(DevicesDiagnostics, DeviceDiagnosicAdmin)
admin.site.register(OnecContracts, OnecContractsAdmin)
admin.site.register(OnecContacts, OnecContactsAdmin)
admin.site.register(InfoServObj, InfoServObjAdmin)
admin.site.register(InfoServTarifs, InfoServTarifsAdmin)
admin.site.register(InfoServTarifClient, InfoServTarifClientAdmin)
admin.site.register(CellOperator, CellOperatorAdmin)
#admin.site.register(Billing, BillingAdmin)
admin.site.register(TransferClient, TransferClientAdmin)
